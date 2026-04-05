"""CRUD endpoints for companies."""

import io
import uuid
from datetime import datetime, date
from decimal import Decimal, InvalidOperation
from typing import Optional, List

from fastapi import APIRouter, Depends, HTTPException, Query, UploadFile, File
from sqlalchemy import select, func, or_, and_, cast, String
from sqlalchemy.ext.asyncio import AsyncSession

from ..database import get_db
from ..models import (
    Company, Interaction,
    CompanyCreate, CompanyUpdate, CompanyResponse, ImportResult,
)

router = APIRouter(prefix="/api/companies", tags=["companies"])

COLUMN_MAP = {
    "empresa": "razao_social",
    "razão social": "razao_social",
    "razao social": "razao_social",
    "cnpj": "cnpj",
    "cnpj completo": "cnpj",
    "uf": "uf",
    "score vf": "score_vf",
    "score": "score_vf",
    "prioridade": "prioridade",
    "faixa de valor": "faixa_valor",
    "faixa valor": "faixa_valor",
    "valor aberto": "valor_aberto",
    "valor aberto (r$)": "valor_aberto",
    "total consolidado": "valor_aberto",
    "qtd inscrições": "qtd_inscricoes",
    "qtd inscricoes": "qtd_inscricoes",
    "qtd. inscrições": "qtd_inscricoes",
    "quantidade de inscrições": "qtd_inscricoes",
    "anos pgfn": "anos_pgfn",
    "anos na pgfn": "anos_pgfn",
    "ano última inscrição": "ano_ult_inscricao",
    "ano últ. inscrição": "ano_ult_inscricao",
    "ano ult inscricao": "ano_ult_inscricao",
    "situação processual": "situacao_processual",
    "situacao processual": "situacao_processual",
    "situações presentes": "situacao_processual",
    "receita principal": "receita_principal",
    "simples nacional": "simples_nacional",
    "seguradora elegível": "seguradora_elegivel",
    "seguradora elegivel": "seguradora_elegivel",
    "seguradora": "seguradora_elegivel",
    "estágio pipeline": "estagio_pipeline",
    "estagio pipeline": "estagio_pipeline",
    "responsável v&f": "responsavel",
    "responsavel": "responsavel",
    "próximo follow-up": "proximo_followup",
    "proximo followup": "proximo_followup",
    "contato (nome)": "decisor_nome",
    "contato": "decisor_nome",
    "telefone": "decisor_telefone",
    "e-mail": "decisor_email",
    "email": "decisor_email",
    "porte": "porte",
}


def _safe_decimal(val) -> Optional[Decimal]:
    if val is None:
        return None
    try:
        if isinstance(val, str):
            val = val.replace(".", "").replace(",", ".").strip()
            if val in ("", "-"):
                return None
        return Decimal(str(val))
    except (InvalidOperation, ValueError):
        return None


def _safe_int(val) -> Optional[int]:
    if val is None:
        return None
    try:
        return int(float(str(val)))
    except (ValueError, TypeError):
        return None


def _safe_date(val):
    if val is None:
        return None
    if isinstance(val, date):
        return val
    if isinstance(val, datetime):
        return val.date()
    try:
        s = str(val).strip()
        if not s or s in ("", "-", "None"):
            return None
        # Try common formats
        for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y"):
            try:
                return datetime.strptime(s, fmt).date()
            except ValueError:
                continue
        return None
    except (ValueError, TypeError):
        return None


def _safe_bool(val) -> Optional[bool]:
    if val is None:
        return None
    if isinstance(val, bool):
        return val
    s = str(val).strip().lower()
    if s in ("sim", "s", "yes", "y", "true", "1"):
        return True
    if s in ("não", "nao", "n", "no", "false", "0"):
        return False
    return None


@router.get("", response_model=List[CompanyResponse])
async def list_companies(
    status: str = Query("active"),
    frente: Optional[int] = None,
    responsavel: Optional[str] = None,
    estagio_pipeline: Optional[str] = None,
    seguradora_elegivel: Optional[str] = None,
    uf: Optional[str] = None,
    score_min: Optional[float] = None,
    score_max: Optional[float] = None,
    faixa_valor: Optional[str] = None,
    situacao_processual: Optional[str] = None,
    simples_nacional: Optional[bool] = None,
    enrichment_status: Optional[str] = None,
    search: Optional[str] = None,
    limit: int = Query(500, le=5000),
    offset: int = Query(0, ge=0),
    db: AsyncSession = Depends(get_db),
):
    q = select(Company).where(Company.status == status)

    if frente is not None:
        q = q.where(Company.frente == frente)
    if responsavel:
        q = q.where(Company.responsavel == responsavel)
    if estagio_pipeline:
        q = q.where(Company.estagio_pipeline == estagio_pipeline)
    if seguradora_elegivel:
        q = q.where(Company.seguradora_elegivel == seguradora_elegivel)
    if uf:
        q = q.where(Company.uf == uf)
    if score_min is not None:
        q = q.where(Company.score_vf >= Decimal(str(score_min)))
    if score_max is not None:
        q = q.where(Company.score_vf <= Decimal(str(score_max)))
    if faixa_valor:
        q = q.where(Company.faixa_valor == faixa_valor)
    if situacao_processual:
        q = q.where(Company.situacao_processual == situacao_processual)
    if simples_nacional is not None:
        q = q.where(Company.simples_nacional == simples_nacional)
    if enrichment_status:
        q = q.where(Company.enrichment_status == enrichment_status)
    if search:
        pattern = f"%{search}%"
        q = q.where(
            or_(
                Company.razao_social.ilike(pattern),
                Company.cnpj.ilike(pattern),
            )
        )

    q = q.order_by(Company.score_vf.desc().nullslast()).offset(offset).limit(limit)
    result = await db.execute(q)
    return result.scalars().all()


@router.post("", response_model=CompanyResponse, status_code=201)
async def create_company(
    company: CompanyCreate,
    db: AsyncSession = Depends(get_db),
):
    existing = await db.execute(
        select(Company).where(Company.cnpj == company.cnpj)
    )
    if existing.scalar_one_or_none():
        raise HTTPException(400, f"CNPJ {company.cnpj} already exists")

    obj = Company(**company.model_dump())
    obj.data_entrada_estagio = datetime.utcnow()
    db.add(obj)
    await db.flush()
    await db.refresh(obj)
    return obj


@router.get("/{company_id}", response_model=CompanyResponse)
async def get_company(
    company_id: uuid.UUID,
    db: AsyncSession = Depends(get_db),
):
    result = await db.execute(select(Company).where(Company.id == company_id))
    company = result.scalar_one_or_none()
    if not company:
        raise HTTPException(404, "Company not found")
    return company


@router.put("/{company_id}", response_model=CompanyResponse)
async def update_company(
    company_id: uuid.UUID,
    update: CompanyUpdate,
    force: bool = Query(False, description="Force stage change bypassing validation"),
    db: AsyncSession = Depends(get_db),
):
    from ..validation import validate_transition

    result = await db.execute(select(Company).where(Company.id == company_id))
    company = result.scalar_one_or_none()
    if not company:
        raise HTTPException(404, "Company not found")

    data = update.model_dump(exclude_unset=True)

    # Track stage changes with validation
    if "estagio_pipeline" in data and data["estagio_pipeline"] != company.estagio_pipeline:
        old_stage = company.estagio_pipeline
        new_stage = data["estagio_pipeline"]

        # Validate transition unless forced
        if not force and old_stage:
            valid, missing = await validate_transition(db, company_id, old_stage, new_stage)
            if not valid:
                raise HTTPException(
                    422,
                    detail={
                        "message": f"Transição {old_stage} → {new_stage} bloqueada",
                        "missing": missing,
                        "hint": "Use ?force=true para forçar (será registrado como override)",
                    }
                )

        company.data_entrada_estagio = datetime.utcnow()
        canal = "Override" if force else "Interno"
        resumo = f"Estágio alterado: {old_stage} → {new_stage}"
        if force:
            resumo += " (OVERRIDE — validação ignorada)"

        interaction = Interaction(
            company_id=company_id,
            responsavel=data.get("updated_by", "Sistema"),
            canal=canal,
            resumo=resumo,
            estagio_anterior=old_stage,
            estagio_novo=new_stage,
        )
        db.add(interaction)

    for key, value in data.items():
        setattr(company, key, value)

    await db.flush()
    await db.refresh(company)
    return company


@router.delete("/{company_id}")
async def archive_company(
    company_id: uuid.UUID,
    db: AsyncSession = Depends(get_db),
):
    result = await db.execute(select(Company).where(Company.id == company_id))
    company = result.scalar_one_or_none()
    if not company:
        raise HTTPException(404, "Company not found")
    company.status = "archived"
    await db.flush()
    return {"message": "Company archived"}


@router.post("/import", response_model=ImportResult)
async def import_companies(
    file: UploadFile = File(...),
    db: AsyncSession = Depends(get_db),
):
    if not file.filename.endswith((".xlsx", ".xls", ".csv")):
        raise HTTPException(400, "Only .xlsx, .xls, and .csv files are supported")

    try:
        import openpyxl
    except ImportError:
        raise HTTPException(500, "openpyxl not installed")

    content = await file.read()
    wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)

    result = ImportResult()
    errors = []

    for sheet in wb.sheetnames:
        # Skip non-data sheets
        sheet_lower = sheet.lower()
        if "painel" in sheet_lower or "resumo" in sheet_lower:
            continue

        # Detect frente/seguradora from sheet name
        sheet_frente = None
        sheet_seguradora = None
        if "sancor" in sheet_lower:
            sheet_frente = 1
            sheet_seguradora = "Sancor"
        elif "berkley" in sheet_lower:
            sheet_frente = 2
            sheet_seguradora = "Berkley"
        elif "zurich" in sheet_lower or "swiss" in sheet_lower or "chubb" in sheet_lower:
            sheet_frente = 2
            sheet_seguradora = "Zurich/Swiss/Chubb"
        elif "completa" in sheet_lower or "base" in sheet_lower:
            sheet_frente = 1
            sheet_seguradora = "Sancor"

        ws = wb[sheet]
        header_row = None
        header_map = {}

        # Find header row (contains "Empresa" or "CNPJ")
        for row_idx, row in enumerate(ws.iter_rows(max_row=30, values_only=False), 1):
            values = [str(c.value).strip().lower() if c.value else "" for c in row]
            has_key = any(v in ("empresa", "cnpj", "cnpj completo", "razão social", "razao social") for v in values)
            if has_key:
                header_row = row_idx
                for col_idx, val in enumerate(values):
                    if val in COLUMN_MAP:
                        header_map[col_idx] = COLUMN_MAP[val]
                break

        if not header_row or "cnpj" not in header_map.values():
            continue

        cnpj_col = next(k for k, v in header_map.items() if v == "cnpj")

        for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
            if not row or not row[cnpj_col]:
                continue

            cnpj_raw = str(row[cnpj_col]).strip()
            if not cnpj_raw or cnpj_raw == "None":
                continue

            # Format CNPJ
            digits = "".join(c for c in cnpj_raw if c.isdigit())
            if len(digits) == 14:
                cnpj = f"{digits[:2]}.{digits[2:5]}.{digits[5:8]}/{digits[8:12]}-{digits[12:]}"
            elif len(digits) > 0:
                cnpj = cnpj_raw
            else:
                continue

            try:
                # Parse row into field dict
                row_data = {}
                for col_idx, field_name in header_map.items():
                    if field_name == "cnpj":
                        continue
                    val = row[col_idx] if col_idx < len(row) else None
                    if val is None:
                        continue
                    if field_name in ("score_vf", "valor_aberto", "anos_pgfn"):
                        row_data[field_name] = _safe_decimal(val)
                    elif field_name in ("qtd_inscricoes", "ano_ult_inscricao"):
                        row_data[field_name] = _safe_int(val)
                    elif field_name == "simples_nacional":
                        row_data[field_name] = _safe_bool(val)
                    elif field_name in ("proximo_followup", "data_proposta", "data_emissao", "data_balanco"):
                        row_data[field_name] = _safe_date(val)
                    else:
                        row_data[field_name] = str(val).strip() if val else None

                # Check if CNPJ already exists — consolidate inscriptions
                existing_q = await db.execute(select(Company).where(Company.cnpj == cnpj))
                existing = existing_q.scalar_one_or_none()

                row_valor = row_data.get("valor_aberto") or Decimal("0")

                if existing:
                    # Consolidate: sum valor_aberto, increment qtd_inscricoes
                    existing.valor_aberto = (existing.valor_aberto or Decimal("0")) + row_valor
                    existing.qtd_inscricoes = (existing.qtd_inscricoes or 0) + 1
                    # Keep highest score
                    new_score = row_data.get("score_vf")
                    if new_score and (not existing.score_vf or new_score > existing.score_vf):
                        existing.score_vf = new_score
                    # Merge situacao_processual (append if different)
                    new_sit = row_data.get("situacao_processual")
                    if new_sit and existing.situacao_processual and new_sit not in existing.situacao_processual:
                        existing.situacao_processual = existing.situacao_processual + " | " + new_sit
                    elif new_sit and not existing.situacao_processual:
                        existing.situacao_processual = new_sit
                    # Fill empty fields from new row (decisor, telefone, email, etc)
                    for field in ("decisor_nome", "decisor_telefone", "decisor_email", "porte",
                                  "receita_principal", "responsavel"):
                        new_val = row_data.get(field)
                        if new_val and not getattr(existing, field, None):
                            setattr(existing, field, new_val)
                    result.duplicatas += 1  # count as consolidated
                else:
                    # New company
                    obj = Company(cnpj=cnpj, **row_data)
                    obj.status = "active"
                    obj.enrichment_status = "pgfn_only"
                    if not obj.estagio_pipeline:
                        obj.estagio_pipeline = "Base PGFN"
                    obj.data_entrada_estagio = datetime.utcnow()
                    if not obj.qtd_inscricoes:
                        obj.qtd_inscricoes = 1
                    if sheet_frente and not obj.frente:
                        obj.frente = sheet_frente
                    if sheet_seguradora and not obj.seguradora_elegivel:
                        obj.seguradora_elegivel = sheet_seguradora
                    db.add(obj)
                    result.importadas += 1
            except Exception as e:
                result.erros += 1
                errors.append(f"CNPJ {cnpj}: {str(e)}")

    result.detalhes_erros = errors[:50]
    await db.flush()
    return result
